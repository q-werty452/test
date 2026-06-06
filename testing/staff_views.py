"""
Портал сотрудников IT COLLEGE:
  - Вход/выход по логину Django
  - Таблица всех абитуриентов с баллами
  - Детальная страница абитуриента (ответы по каждому вопросу)
  - Печать PDF-отчёта (print-ready HTML → браузерная печать)
  - Экспорт всех результатов в Excel (.xlsx)
"""
import io
from datetime import datetime

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Abiturient, TestResult, AbiturientAnswer, Subject, TestVariant, SiteSettings, Question

# URL для редиректа при недостатке прав
_LOGIN_URL = 'staff_login'


# ─────────────────────────────────────────
#  АВТОРИЗАЦИЯ
# ─────────────────────────────────────────

def staff_login_view(request):
    """Страница входа для сотрудников."""
    if request.user.is_authenticated:
        return redirect('staff_dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_staff or user.is_superuser:
                login(request, user)
                return redirect(request.GET.get('next', 'staff_dashboard'))
            else:
                error = 'У вашей учётной записи нет прав доступа к порталу сотрудников.'
        else:
            error = 'Неверный логин или пароль. Попробуйте ещё раз.'

    return render(request, 'testing/staff/login.html', {'error': error})


def staff_logout_view(request):
    logout(request)
    return redirect('staff_login')


# ─────────────────────────────────────────
#  ДАШБОРД — список всех абитуриентов
# ─────────────────────────────────────────

@login_required(login_url=_LOGIN_URL)
def staff_dashboard_view(request):
    """Главная страница портала сотрудников."""
    search      = request.GET.get('search', '').strip()
    grade_f     = request.GET.get('grade', '')
    variant_f   = request.GET.get('variant', '')
    completed_f = request.GET.get('completed', '')

    qs = (
        Abiturient.objects
        .select_related('result', 'test_variant')
        .order_by('-started_at')
    )

    if search:
        qs = qs.filter(
            Q(last_name__icontains=search)              |
            Q(first_name__icontains=search)             |
            Q(middle_name__icontains=search)            |
            Q(pin__icontains=search)                    |
            Q(birth_certificate__icontains=search)      |
            Q(school__icontains=search)
        )
    if grade_f:
        qs = qs.filter(grade=grade_f)
    if variant_f:
        qs = qs.filter(test_variant_id=variant_f)
    if completed_f == '1':
        qs = qs.filter(is_completed=True)
    elif completed_f == '0':
        qs = qs.filter(is_completed=False)

    # Сводная статистика
    total_all    = Abiturient.objects.count()
    completed_all= Abiturient.objects.filter(is_completed=True).count()
    avg_obj      = TestResult.objects.aggregate(avg=Avg('total_score'))
    avg_score    = round(avg_obj['avg'], 1) if avg_obj['avg'] else 0

    site = SiteSettings.get()

    context = {
        'abiturients':    qs,
        'total_all':      total_all,
        'completed_all':  completed_all,
        'avg_score':      avg_score,
        'variants':       TestVariant.objects.all(),
        'grade_choices':  Abiturient.GRADE_CHOICES,
        'search':         search,
        'grade_f':        grade_f,
        'variant_f':      variant_f,
        'completed_f':    completed_f,
        'site':           site,
    }
    return render(request, 'testing/staff/dashboard.html', context)


# ─────────────────────────────────────────
#  ДЕТАЛЬНАЯ СТРАНИЦА АБИТУРИЕНТА
# ─────────────────────────────────────────

@login_required(login_url=_LOGIN_URL)
def staff_student_detail_view(request, pk):
    """Страница абитуриента: личные данные + все ответы по предметам."""
    abiturient = get_object_or_404(
        Abiturient.objects.select_related('result', 'test_variant'), pk=pk
    )

    answers = (
        AbiturientAnswer.objects
        .filter(abiturient=abiturient)
        .select_related('question__subject', 'selected_option')
        .prefetch_related('question__options')
        .order_by('question__subject__order', 'question__order')
    )

    subjects_data = _group_answers_by_subject(answers)
    english_max, math_max, russian_max, max_total = _compute_max_scores(subjects_data)

    context = {
        'abiturient':    abiturient,
        'subjects_data': subjects_data,
        'english_max':   english_max,
        'math_max':      math_max,
        'russian_max':   russian_max,
        'max_total':     max_total,
    }
    return render(request, 'testing/staff/student_detail.html', context)


# ─────────────────────────────────────────
#  ПЕЧАТЬ / PDF-ОТЧЁТ
# ─────────────────────────────────────────

@login_required(login_url=_LOGIN_URL)
def staff_print_report_view(request, pk):
    """
    Страница-ведомость, оптимизированная для печати / сохранения как PDF.
    Открывается в новой вкладке → Ctrl+P → «Сохранить как PDF».

    Когда получим стандартный бланк колледжа, здесь заменим шаблон
    и добавим наложение данных поверх загруженного PDF.
    """
    abiturient = get_object_or_404(
        Abiturient.objects.select_related('result', 'test_variant'), pk=pk
    )

    answers = (
        AbiturientAnswer.objects
        .filter(abiturient=abiturient)
        .select_related('question__subject', 'selected_option')
        .prefetch_related('question__options')
        .order_by('question__subject__order', 'question__order')
    )

    subjects_data = _group_answers_by_subject(answers)
    english_max, math_max, russian_max, max_total = _compute_max_scores(subjects_data)

    now = timezone.now()
    course = 1 if abiturient.grade == '9' else 2
    year1 = now.year if now.month < 9 else now.year + 1
    academic_year = f'{year1}–{year1 + 1}'

    context = {
        'abiturient':    abiturient,
        'subjects_data': subjects_data,
        'now':           now,
        'course':        course,
        'academic_year': academic_year,
        'english_max':   english_max,
        'math_max':      math_max,
        'russian_max':   russian_max,
        'max_total':     max_total,
    }
    return render(request, 'testing/staff/print_report.html', context)


# ─────────────────────────────────────────
#  ЭКСПОРТ В EXCEL
# ─────────────────────────────────────────

@login_required(login_url=_LOGIN_URL)
def export_excel_view(request):
    """
    Генерирует и скачивает Excel-файл со всеми абитуриентами и их результатами.
    Цветовая кодировка: зелёный ≥ 42, жёлтый ≥ 30, красный < 30.
    """
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Результаты тестирования'

    # ── Стили ──
    WHITE   = 'FFFFFF'
    DARK    = '060C1D'
    BLUE    = '1D4ED8'
    GRAY    = '64748B'
    LIGHT   = 'EFF2FF'

    def _font(bold=False, size=11, color=DARK):
        return Font(name='Calibri', bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def _align(h='center', wrap=False):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

    _side   = Side(style='thin', color='DDE5FF')
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    # ── Шапка листа ──
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'IT COLLEGE — Результаты вступительного тестирования'
    ws['A1'].font      = _font(bold=True, size=14)
    ws['A1'].alignment = _align()
    ws['A1'].fill      = _fill('EFF2FF')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:Q2')
    ws['A2'] = f'Сформировано: {datetime.now().strftime("%d.%m.%Y  %H:%M")}'
    ws['A2'].font      = _font(size=10, color=GRAY)
    ws['A2'].alignment = _align()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # ── Заголовки столбцов ──
    columns = [
        ('№',              5),
        ('Фамилия',       18),
        ('Имя',           14),
        ('Отчество',      16),
        ('Дата рождения', 14),
        ('Школа / Заведение', 30),
        ('Класс',          8),
        ('ПИН',           16),
        ('Вариант',       12),
        ('Дата теста',    13),
        ('Начало',        10),
        ('Конец',         10),
        ('Английский', 13),
        ('Математика', 13),
        ('Русский яз.', 13),
        ('Итого', 11),
        ('%',              8),
    ]

    for col_idx, (title, width) in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=title)
        cell.font      = _font(bold=True, color=WHITE)
        cell.fill      = _fill(BLUE)
        cell.alignment = _align(wrap=True)
        cell.border    = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 38
    ws.freeze_panes = 'A5'

    # ── Данные ──
    abiturients = list(
        Abiturient.objects
        .select_related('result', 'test_variant')
        .order_by('last_name', 'first_name', 'middle_name')
    )

    for row_idx, ab in enumerate(abiturients, 1):
        r      = row_idx + 4
        result = getattr(ab, 'result', None)

        eng   = result.english_score if result else '—'
        mat   = result.math_score    if result else '—'
        rus   = result.russian_score if result else '—'
        total = result.total_score   if result else '—'
        max_q = Question.objects.filter(variant=ab.test_variant).count() if ab.test_variant else 0
        pct   = f'{round(result.total_score / max_q * 100)}%' if (result and max_q) else '—'

        row_values = [
            row_idx,
            ab.last_name, ab.first_name, ab.middle_name,
            ab.birth_date.strftime('%d.%m.%Y') if ab.birth_date else '—',
            ab.school,
            ab.get_grade_display(),
            ab.doc_number,
            ab.test_variant.name if ab.test_variant else '—',
            ab.started_at.astimezone().strftime('%d.%m.%Y') if ab.started_at else '—',
            ab.started_at.astimezone().strftime('%H:%M')     if ab.started_at else '—',
            ab.finished_at.astimezone().strftime('%H:%M')    if ab.finished_at else '—',
            eng, mat, rus, total, pct,
        ]

        left_cols = {2, 3, 4, 6, 8}   # ФИО, Школа, ПИН — выравнивание по левому краю
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border    = _border
            cell.alignment = _align('left' if col_idx in left_cols else 'center')
            cell.font      = _font(size=10)

        # Цвет итоговой ячейки
        if result:
            score_cell      = ws.cell(row=r, column=16)
            score_cell.font = _font(bold=True, size=10)
            if result.total_score >= 42:
                score_cell.fill = _fill('DCFCE7')   # зелёный
                score_cell.font = _font(bold=True, size=10, color='166534')
            elif result.total_score >= 30:
                score_cell.fill = _fill('FEF9C3')   # жёлтый
                score_cell.font = _font(bold=True, size=10, color='854D0E')
            else:
                score_cell.fill = _fill('FEE2E2')   # красный
                score_cell.font = _font(bold=True, size=10, color='991B1B')

        # Чередование строк
        if row_idx % 2 == 0:
            for col_idx in range(1, 18):
                cell = ws.cell(row=r, column=col_idx)
                if cell.fill.fgColor.rgb in ('00000000', '00FFFFFF'):
                    cell.fill = _fill('F8FAFF')

        ws.row_dimensions[r].height = 20

    # ── Итоговая строка ──
    total_count = len(abiturients)
    last_r = total_count + 5
    ws.merge_cells(f'A{last_r}:L{last_r}')
    ws.cell(row=last_r, column=1, value=f'Всего абитуриентов: {total_count}').font = _font(bold=True)
    ws.cell(row=last_r, column=1).alignment = _align('left')

    # ── Отправляем файл ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts       = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'itcollege_results_{ts}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────
#  ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ
# ─────────────────────────────────────────

@login_required(login_url=_LOGIN_URL)
def exam_settings_view(request):
    """Включение/выключение экзамена и управление кодом доступа."""
    if request.method != 'POST':
        return redirect('staff_dashboard')

    site = SiteSettings.get()
    action = request.POST.get('action')

    if action == 'toggle':
        site.exam_is_active = not site.exam_is_active
        site.save(update_fields=['exam_is_active', 'updated_at'])
        status = 'ОТКРЫТО' if site.exam_is_active else 'ЗАКРЫТО'
        messages.success(request, f'Тестирование {status}.')

    elif action == 'set_code':
        code = request.POST.get('access_code', '').strip()
        site.access_code = code.upper() if code else ''
        site.save(update_fields=['access_code', 'updated_at'])
        if code:
            messages.success(request, f'Код доступа установлен: {site.access_code}')
        else:
            messages.success(request, 'Код доступа отключён — регистрация без кода.')

    return redirect('staff_dashboard')


def _compute_max_scores(subjects_data):
    """Возвращает (english_max, math_max, russian_max, total_max) из данных subjects_data."""
    english_max = math_max = russian_max = 0
    for g in subjects_data:
        name_lower = g['subject'].name.lower()
        if 'англ' in name_lower or 'english' in name_lower:
            english_max = g['total']
        elif 'мат' in name_lower or 'math' in name_lower:
            math_max = g['total']
        elif 'рус' in name_lower or 'russian' in name_lower:
            russian_max = g['total']
    return english_max, math_max, russian_max, english_max + math_max + russian_max


def _group_answers_by_subject(answers_qs):
    """Группирует QuerySet ответов по предмету, возвращает упорядоченный список."""
    groups = {}
    for answer in answers_qs:
        subj = answer.question.subject
        if subj.id not in groups:
            groups[subj.id] = {
                'subject': subj,
                'answers': [],
                'correct': 0,
                'total':   0,
            }
        # Находим правильный вариант для этого вопроса
        correct_opt = answer.question.options.filter(is_correct=True).first()
        groups[subj.id]['answers'].append({
            'answer':      answer,
            'correct_opt': correct_opt,
        })
        groups[subj.id]['total'] += 1
        if answer.is_correct:
            groups[subj.id]['correct'] += 1

    return sorted(groups.values(), key=lambda g: g['subject'].order)
